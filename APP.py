import streamlit as st
import pandas as pd
import numpy as np
import re
import io
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# Configuración de la página
st.set_page_config(
    page_title="Auditoría de Trabajo Suplementario",
    page_icon="📋",
    layout="wide"
)

# Estilos CSS definitivos para forzar el diseño plano sin cuadros de fondo al cargar
st.markdown("""
    <div style="background-color:#1F4E79; padding:10px 20px; border-radius:8px; text-align:center; margin-bottom:15px;">
        <h2 style="color:white; margin:0; font-size:20px; font-weight:600;">SaberBot Auditoría de Trabajo Suplementario</h2>
    </div>
    <style>
        /* Ajustar el contenedor general del cargador */
        [data-testid="stFileUploader"] {
            padding: 0px;
        }
        /* Transforma la zona de arrastre en un recuadro limpio y plano */
        [data-testid="stFileUploaderDropzone"] {
            padding: 8px 12px !important;
            border: 1px dashed #bfc7d2 !important;
            border-radius: 4px !important;
            background-color: #f8f9fa !important;
            display: flex !important;
            flex-direction: row !important;
            align-items: center !important;
            justify-content: flex-start !important;
            gap: 12px !important;
            min-height: 45px !important;
        }
        /* Ocultar textos nativos de Streamlit ("Drag and drop", tamaño máximo, etc.) */
        [data-testid="stFileUploaderDropzone"] section,
        [data-testid="stFileUploaderDropzone"] small,
        [data-testid="stFileUploaderDropzone"] svg {
            display: none !important;
        }
        /* Inyectar de forma simulada el botón gris clásico de "Seleccionar archivo" */
        [data-testid="stFileUploaderDropzone"]::before {
            content: "Seleccionar archivo" !important;
            display: inline-block !important;
            background-color: #f0f0f0 !important;
            color: #333333 !important;
            border: 1px solid #767676 !important;
            border-radius: 2px !important;
            padding: 3px 8px !important;
            font-family: Arial, sans-serif !important;
            font-size: 13px !important;
            cursor: pointer !important;
        }
        /* Ocultar el botón original de Streamlit */
        [data-testid="stFileUploaderDropzone"] button {
            display: none !important;
        }
        
        /* ── REMOCIÓN DE CAJAS EXTRA Y CONTENEDORES OSCUROS DE CARGA ── */
        /* Forzar a que la sección interna que almacena los archivos no dibuje un bloque blanco flotante */
        [data-testid="stFileUploaderDropzone"] > div {
            background-color: transparent !important;
            border: none !important;
            padding: 0px !important;
            margin: 0px !important;
            box-shadow: none !important;
        }
        /* Eliminar la estructura del widget de archivo que dibuja el fondo oscuro de fondo */
        [data-testid="stFileUploaderRow"], 
        [data-testid="stFileUploaderRow"] > div {
            background-color: transparent !important;
            border: none !important;
            box-shadow: none !important;
            padding: 0px !important;
            margin: 0px !important;
        }
        /* Ocultar el ícono negro de tipo de archivo, metadatos en MB y el botón nativo de basura */
        [data-testid="stFileUploaderRow"] svg,
        [data-testid="stFileUploaderRow"] button,
        [data-testid="stFileUploaderRow"] data,
        [data-testid="stFileUploaderRow"] small {
            display: none !important;
        }
        /* Ajustar el texto plano para que se sitúe ordenadamente al lado del botón generado */
        [data-testid="stFileUploaderRow"] span {
            font-family: Arial, sans-serif !important;
            font-size: 14px !important;
            color: #333333 !important;
            font-weight: normal !important;
            padding-left: 5px !important;
        }
    </style>
""", unsafe_allow_html=True)


# ── Configuración Inicial Interna ─────────────────────────────────────────
MAPA_ausencias = {
    "ninguno":                   None,
    "ausencia":                  "A",
    "ausencia sin justa causa sd": "FSJ",
    "enfermedad comun":          "EC",
    "accidente de trabajo":      "AT",
    "descanso":                  "DES",
    "dia de la familia":         "DF",
    "licencia no remunerada":    "LNR",
    "licencia remunerada":       "LR",
    "parada de salario":         "PS",
    "suspensiones":              "S",
    "descanso compensatorio":    "C",
    "festivo compensatorio":     "FC",
    "vacaciones":                "V",
    "incapacidad":               "INC",
    "licencia de maternidad":    "LM",
    "remunerada":                "LR",
    "no remunerada":             "LNR",
    "licencia por luto":         "LT",
    "calamidad doméstica":       "CD",
    "falta sin justificar":      "FSJ",
    "suspensión":                "S",
    "CCCO":                      "CCCO",
    "retiros":                   "R",
    "perdida dominical":         "PD",
    "p":                         "P"   
}

AUSENCIAS_ROJO_SET = {
    "A", "INC", "AT", "LM", "LR", "LNR", "DF", "CD",
    "S", "FC", "DES", "PS", "FSJ", "PD"
}
AUSENCIAS_AMARILLO_SET = {"P"}
MESES_ES = {1:"ene",2:"feb",3:"mar",4:"abr",5:"may",6:"jun",7:"jul",8:"ago",9:"sep",10:"oct",11:"nov",12:"dic"}

PERIODOS = {
    "Marzo":      (pd.Timestamp(2026, 2, 14), pd.Timestamp(2026, 3, 20)),
    "Abril":      (pd.Timestamp(2026, 3, 21), pd.Timestamp(2026, 4, 17)),
    "Mayo":       (pd.Timestamp(2026, 4, 18), pd.Timestamp(2026, 5, 15)),
    "Junio":      (pd.Timestamp(2026, 5, 16), pd.Timestamp(2026, 6, 12)),
    "Julio":      (pd.Timestamp(2026, 6, 13), pd.Timestamp(2026, 7, 17)),
    "Agosto":     (pd.Timestamp(2026, 7, 18), pd.Timestamp(2026, 8, 14)),
    "Septiembre": (pd.Timestamp(2026, 8, 15), pd.Timestamp(2026, 9, 18)),
    "Octubre":    (pd.Timestamp(2026, 9, 19), pd.Timestamp(2026, 10, 16)),
    "Noviembre":  (pd.Timestamp(2026, 10, 17), pd.Timestamp(2026, 11, 13)),
    "Diciembre":  (pd.Timestamp(2026, 11, 14), pd.Timestamp(2026, 12, 31))
}

MAPA_CONCEPTOS = {
    "HT Normales":                      "ht normales",
    "Recargo Nocturno 0.35%":           "recargo nocturno 0.35%",
    "Recargo Dominical Compensado":     "recargo dominical compensado",
    "Recargo Dominical No Compensado":  "recargo dominical no compensado",
    "Horas Extras Diurnas 1.25%":       "horas extras diurnas 1.25%",
    "Recargo Festivo":                  "recargo festivo",
    "Hora Extra Diurna Dom/Fest":       "hora extra diurna dom/fest",
    "Recargo Festivo Adicional":        "recargo festivo adicional",
}

COLORES_LETRAS = {
    "A":   {"bg": "FF0000", "fg": "FFFFFF"},
    "S":   {"bg": "FF0000", "fg": "FFFFFF"},
    "INC": {"bg": "FF0000", "fg": "FFFFFF"},
    "C":   {"bg": "F4B942", "fg": "000000"},
    "P":   {"bg": "C00000", "fg": "FFFFFF"},
    "LIC": {"bg": "FFC000", "fg": "000000"},
    "DIA": {"bg": "92D050", "fg": "000000"},
}

def hhmm_a_decimal(texto):
    if pd.isna(texto): return None
    t = str(texto).strip()
    if t in ("", "0", "nan", "None"): return None
    m = re.match(r'^(\d+):(\d{2})$', t)
    if m:
        val = int(m.group(1)) + int(m.group(2)) / 60.0
        return round(val, 2) if val > 0 else None
    try:
        val = float(t)
        return round(val, 2) if val > 0 else None
    except ValueError: return None

def num_limpio(val):
    if pd.isna(val): return None
    try:
        v = float(val)
        return round(v, 2) if v > 0 else None
    except (ValueError, TypeError): return None

def resolver_ausencia(texto_ausent):
    if pd.isna(texto_ausent): return None
    t = str(texto_ausent).strip()
    if not t or t.lower() in ('nan', 'none'): return None
    clave = t.lower()
    if clave in MAPA_ausencias: return MAPA_ausencias[clave]
    for k, v in MAPA_ausencias.items():
        if k and k in clave: return v
    return t

def celda(valor, ausent_raw):
    v = valor
    a = resolver_ausencia(ausent_raw)
    if a == 'C': return 'C'
    if v is not None and a is not None:
        v_str = str(int(v)) if v % 1 == 0 else str(round(v, 2))
        return f'{v_str} {a}'
    if v is not None: return int(v) if v % 1 == 0 else round(v, 2)
    return a

def limpiar_id_a_texto(valor):
    # Si lo que recibe es una columna de Pandas (Series)
    if isinstance(valor, pd.Series):
        temp = pd.to_numeric(valor, errors='coerce')
        return temp.fillna(0).astype(int).astype(str).replace('0', '')
    
    # Si lo que recibe es una sola celda/valor suelto (Scalar)
    if pd.isna(valor): 
        return ""
    s = str(valor).strip()
    if s.lower() in ("nan", "none", "0", ""): 
        return ""
    if s.endswith(".0"): 
        s = s[:-2]
    try:
        return str(int(float(s)))
    except:
        return s

# ── Paso 1: Carga de Archivos ─────────────────────────────────────────────
st.header("📁 1. Carga de Archivos Base")

col_file1, col_file2 = st.columns(2)
with col_file1:
    archivo_cargado = st.file_uploader("Subir plantilla de Asistencia (.xlsx)", type=["xlsx"])
with col_file2:
    archivo_htcc = st.file_uploader("Subir plantilla de Consolidación HTCC (.xlsx)", type=["xlsx"])

if archivo_cargado is not None and archivo_htcc is not None:
    xl = pd.ExcelFile(archivo_cargado)
    hojas_disponibles = xl.sheet_names
    
    xl_htcc = pd.ExcelFile(archivo_htcc)
    hojas_htcc = xl_htcc.sheet_names
    
    col_sheet1, col_sheet2 = st.columns(2)
    with col_sheet1:
        HOJA_ENTRADA = st.selectbox("Seleccione la hoja de marcación:", hojas_disponibles)
    with col_sheet2:
        HOJA_LIBRO3 = st.selectbox("Seleccione la hoja de destino en HTCC:", hojas_htcc)
    
    # ── Paso 2: Parámetros de Fechas ─────────────────────────────────────────
    st.header("📅 2. Parámetros de Filtrado y Fechas")
    fecha_inicio_input = st.date_input("Fecha Inicial de Semanas para Análisis de Compensatorios")
    
    if "procesado" not in st.session_state:
        st.session_state.procesado = False
        st.session_state.output_bytes = None
        st.session_state.htcc_bytes = None
        st.session_state.listado = None
        st.session_state.resumen = None
        st.session_state.df_c = None
        st.session_state.resultados_c = None

    if st.button("🚀 Procesar Reporte e Historial", type="primary"):
        with st.spinner("Procesando datos y estructurando archivos de Excel..."):
            try:
                # Carga de datos origen
                df_origen = pd.read_excel(archivo_cargado, sheet_name=HOJA_ENTRADA)
                df = df_origen.copy()
                
                col_map = {str(c).strip().lower(): c for c in df.columns}
                def buscar_col(patrones):
                    for pat in patrones:
                        for k, orig in col_map.items():
                            if re.search(pat, k, re.IGNORECASE): return orig
                    return None

                col_id     = buscar_col([r'^identificador$'])
                col_fecha  = buscar_col([r'^fecha ori', r'^fecha$'])
                col_dia    = buscar_col([r'^dia$'])
                col_ht     = buscar_col([r'^ht$'])
                col_ausent = buscar_col([r'^ausentismo$'])

                col_rno   = buscar_col([r'recargo nocturno'])
                col_rdc   = buscar_col([r'dominical compensado'])
                col_rdnc  = buscar_col([r'dominical no compensado'])
                col_rf    = buscar_col([r'recargo festivo'])
                col_hed   = buscar_col([r'extras diurnas.*1[\.,\s]?25'])
                col_hedf  = buscar_col([r'extra diurna dom'])
                col_hen   = buscar_col([r'extras nocturnas.*1[\.,\s]?75'])
                col_hendf = buscar_col([r'dominical o festiva nocturna'])

                criticas = {'Identificador':col_id,'Fecha':col_fecha,'Dia':col_dia,'HT':col_ht,'Ausentismo':col_ausent}
                faltantes = [k for k,v in criticas.items() if v is None]
                if faltantes:
                    st.error(f"❌ Columnas críticas no detectadas en la hoja elegida: {faltantes}.")
                    st.stop()

                df['FechaReal'] = pd.to_datetime(df[col_fecha], errors='coerce')
                df['FechaCorta'] = df['FechaReal'].apply(lambda d: f"{d.day}-{MESES_ES[d.month]}" if pd.notna(d) else 'sin-fecha')
                
                def limpiar_dia(texto):
                    if pd.isna(texto): return ''
                    for p in str(texto).strip().split():
                        pl = p.lower()
                        if pl in {'lunes','martes','miércoles','miercoles','jueves','viernes','sábado','sabado','domingo'}:
                            return {'miercoles': 'miércoles', 'sabado': 'sábado'}.get(pl, pl)
                    return str(texto).strip().lower()

                df['dia']      = df[col_dia].apply(limpiar_dia)
                df['_festivo'] = df[col_dia].apply(lambda x: 'festivo' in str(x).strip().lower())

                aus = df[col_ausent]
                df['_HTn']    = [celda(hhmm_a_decimal(h), a) for h, a in zip(df[col_ht], aus)]
                df['_RNn']    = [celda(num_limpio(v), a) for v, a in zip(df[col_rno], aus)]
                df['_RDCn']   = [celda(num_limpio(v), a) for v, a in zip(df[col_rdc], aus)]
                df['_RDNCn']  = [celda(num_limpio(v), a) for v, a in zip(df[col_rdnc], aus)]
                df['_RFn']    = [celda(num_limpio(v), a) for v, a in zip(df[col_rf], aus)]
                df['_HEDn']   = [celda(num_limpio(v), a) for v, a in zip(df[col_hed], aus)]
                df['_HEDFn']  = [celda(num_limpio(v), a) for v, a in zip(df[col_hedf], aus)]
                df['_HENn']   = [celda(num_limpio(v), a) for v, a in zip(df[col_hen], aus)]
                df['_HENDFn'] = [celda(num_limpio(v), a) for v, a in zip(df[col_hendf], aus)]

                CONCEPTOS_COLS = ['_HTn','_RNn','_RDCn','_RDNCn','_RFn','_HEDn','_HEDFn','_HENn','_HENDFn']
                ETIQUETAS = {
                    '_HTn': 'HT Normales', '_RNn': 'Recargo Nocturno 0.35%', '_RDCn': 'Recargo Dominical Compensado',
                    '_RDNCn': 'Recargo Dominical No Compensado', '_RFn': 'Recargo Festivo', '_HEDn': 'Horas Extras Diurnas 1.25%',
                    '_HEDFn': 'Hora Extra Diurna Dom/Fest', '_HENn': 'Horas Extras Nocturnas 1.75%', '_HENDFn': 'Hora Extra Dominical o Festiva Nocturna'
                }
                ORDEN_CONCEPTOS = {v: i+1 for i, v in enumerate(ETIQUETAS.values())}
                COLS_FIJAS = [c for c in [col_id, 'FechaReal', 'FechaCorta', 'dia', '_festivo'] if c in df.columns]

                unpivoted = df[COLS_FIJAS + CONCEPTOS_COLS].melt(id_vars=COLS_FIJAS, value_vars=CONCEPTOS_COLS, var_name='_col', value_name='_valor').reset_index(drop=True)
                unpivoted = unpivoted[~(unpivoted['_valor'].isna() & (unpivoted['_col'] != '_HTn'))].reset_index(drop=True)
                unpivoted['Concepto'] = unpivoted['_col'].map(ETIQUETAS)
                unpivoted['_Orden']   = unpivoted['Concepto'].map(ORDEN_CONCEPTOS).fillna(10).astype(int)
                unpivoted = unpivoted.drop(columns='_col').sort_values([col_id, '_Orden', 'FechaReal']).reset_index(drop=True)

                # Corrección del mapa_festivo
                mapa_fechas = unpivoted[['FechaReal', 'FechaCorta', 'dia', '_festivo']].drop_duplicates(subset=['FechaCorta']).dropna(subset=['FechaReal']).sort_values('FechaReal')
                fechas_unicas  = mapa_fechas['FechaCorta'].tolist()
                mapa_dia       = dict(zip(mapa_fechas['FechaCorta'], mapa_fechas['dia']))
                mapa_festivo   = dict(zip(mapa_fechas['FechaCorta'], mapa_fechas['_festivo']))

                pivotados = []
                for nombre in ETIQUETAS.values():
                    sub = unpivoted[unpivoted['Concepto'] == nombre].sort_values('FechaReal').drop_duplicates(subset=[col_id, 'FechaCorta'], keep='first')
                    piv = sub.pivot(index=[col_id], columns='FechaCorta', values='_valor').reset_index()
                    cols_f = [c for c in piv.columns if c != col_id]
                    piv[cols_f] = piv[cols_f].fillna(0)
                    piv['Concepto'] = nombre
                    piv['_Orden']   = ORDEN_CONCEPTOS[nombre]
                    pivotados.append(piv)

                pivotado = pd.concat(pivotados, ignore_index=True)
                cols_fecha_ok = [f for f in fechas_unicas if f in pivotado.columns]
                pivotado = pivotado[[c for c in pivotado.columns if c not in fechas_unicas] + cols_fecha_ok].fillna(0)

                if 'Nombre' not in df_origen.columns:
                    df_origen['Nombre'] = df_origen['Nombres'].astype(str).str.strip() + " " + df_origen['Apellidos'].astype(str).str.strip()
                mapa_nombres = df_origen[[col_id, 'Nombre']].drop_duplicates(subset=[col_id]).copy()
                mapa_nombres[col_id] = limpiar_id_a_texto(mapa_nombres[col_id])

                final = pivotado.drop(columns=['_Orden'], errors='ignore').rename(columns={col_id: 'Identificador'})
                final['Identificador'] = limpiar_id_a_texto(final['Identificador'])
                final = pd.merge(final, mapa_nombres, left_on='Identificador', right_on=col_id, how='left')
                if col_id != 'Identificador' and col_id in final.columns: final = final.drop(columns=[col_id])

                COLS_INICIO_PRES = [c for c in ['Identificador', 'Nombre', 'Concepto'] if c in final.columns]
                date_cols = [c for c in final.columns if c not in COLS_INICIO_PRES]
                final = final[COLS_INICIO_PRES + date_cols]
                final['_ord_c']  = final['Concepto'].map(ORDEN_CONCEPTOS).fillna(10).astype(int)
                final['_ord_id'] = pd.to_numeric(final['Identificador'], errors='coerce')
                final = final.sort_values(['_ord_id', '_ord_c']).reset_index(drop=True).drop(columns=['_ord_c', '_ord_id'])
                final['Identificador'] = final['Identificador'].astype(str)
                final[date_cols] = final[date_cols].fillna('')

                # Construcción Excel Reporte Horizontal
                output_buffer = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = 'Reporte_Horizontal'

                C_NAVY, C_ROJO, C_VRD_DOM, C_VRD_FES, C_VRD_HDR, C_FES_HDR, C_GRIS, C_NARANJA, C_RDC, C_HE, C_AMBAR, C_AMARILLO = '1F4E79', 'FF0000', 'C6EFCE', '92D050', '538135', '375623', 'D9D9D9', 'FF8C00', 'FFCCCC', 'BDD7EE', 'F4B942', 'FFEB9C'
                thin = Side(style='thin', color='CCCCCC')
                brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

                all_cols = final.columns.tolist()
                for ci, col in enumerate(all_cols, start=1):
                    c = ws.cell(row=1, column=ci, value=col)
                    c.font, c.fill, c.border = Font(name='Arial', bold=True, color='FFFFFF', size=9), PatternFill('solid', fgColor=C_NAVY), brd
                    c.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90 if col in date_cols else 0, wrap_text=(col not in date_cols))
                ws.row_dimensions[1].height = 55

                for ci, col in enumerate(all_cols, start=1):
                    dia, es_fest = mapa_dia.get(col, ''), mapa_festivo.get(col, False)
                    texto_dia = f"{dia} Fest" if es_fest and col in date_cols else dia if col in date_cols else ''
                    bg, fg = C_FES_HDR if es_fest else C_VRD_HDR if 'domingo' in dia else C_GRIS, 'FFFFFF' if (es_fest or 'domingo' in dia) else '000000'
                    c = ws.cell(row=2, column=ci, value=texto_dia)
                    c.font, c.fill, c.alignment, c.border = Font(name='Arial', bold=True, color=fg, size=8), PatternFill('solid', fgColor=bg), Alignment(horizontal='center', vertical='center'), brd
                ws.row_dimensions[2].height = 15

                for ri in range(len(final)):
                    concepto = str(final.iloc[ri]['Concepto']).strip()
                    for ci, col in enumerate(all_cols, start=1):
                        val = final.iloc[ri][col]
                        val_str = str(val).strip() if val is not None else ''
                        abrev = resolver_ausencia(val_str) if col in date_cols and val_str else None
                        
                        if col not in date_cols:
                            if concepto == 'Recargo Dominical No Compensado': bg, fg = C_NARANJA, 'FFFFFF'
                            elif concepto == 'Recargo Dominical Compensado': bg, fg = C_RDC, '000000'
                            elif concepto in {'Horas Extras Diurnas 1.25%', 'Hora Extra Diurna Dom/Fest', 'Horas Extras Nocturnas 1.75%', 'Hora Extra Dominical o Festiva Nocturna'}: bg, fg = C_HE, '000000'
                            else: bg, fg = ('F2F2F2' if ri % 2 == 0 else 'FFFFFF'), '000000'
                        elif col in date_cols and abrev in AUSENCIAS_ROJO_SET: bg, fg = C_ROJO, 'FFFFFF'
                        elif col in date_cols and abrev in AUSENCIAS_AMARILLO_SET: bg, fg = C_AMARILLO, '000000'
                        elif col in date_cols and val_str.lower() == 'c': bg, fg = C_AMBAR, '000000'
                        elif mapa_festivo.get(col, False): bg, fg = C_VRD_FES, '000000'
                        elif 'domingo' in mapa_dia.get(col, ''): bg, fg = C_VRD_DOM, '000000'
                        else: bg, fg = ('FFFFFF' if ri % 2 == 0 else 'F5F5F5'), '000000'

                        c = ws.cell(row=ri + 3, column=ci, value=val)
                        if col in date_cols and isinstance(val, (int, float)): c.number_format = '#,##0.00'
                        c.font, c.fill, c.alignment, c.border = Font(name='Arial', size=9, bold=(col not in date_cols), color=fg), PatternFill('solid', fgColor=bg), Alignment(horizontal='center' if col in date_cols else 'left', vertical='center'), brd
                    ws.row_dimensions[ri + 3].height = 14

                ws.column_dimensions['A'].width, ws.column_dimensions['B'].width = 14, 42
                for ci in range(3, len(all_cols) + 1): ws.column_dimensions[get_column_letter(ci)].width = 6
                ws.freeze_panes, ws.auto_filter.ref = 'D3', f"A2:{get_column_letter(final.shape[1])}{final.shape[0] + 2}"
                wb.save(output_buffer)

                # ── Hoja Ausencias ────────────────────────────────────────────────
                df['Nombre'] = df['Nombres'].astype(str).str.strip() + " " + df['Apellidos'].astype(str).str.strip()
                mask_aus = (df[col_ausent].notna() & (df[col_ausent].astype(str).str.strip().str.lower() == 'ausencia'))
                df_aus = df[mask_aus][[col_id, 'Nombre', 'FechaReal', col_ausent]].copy()
                df_aus['Fecha'] = df_aus['FechaReal'].dt.strftime('%d/%m/%Y')
                df_aus['_ord']  = pd.to_numeric(df_aus[col_id], errors='coerce')
                df_aus = df_aus.sort_values(['_ord', 'FechaReal']).reset_index(drop=True)
                listado = df_aus[[col_id, 'Nombre', 'Fecha', col_ausent]].copy()
                listado.columns = ['Identificador', 'Nombre', 'Fecha', 'Novedad']
                resumen = listado.groupby('Identificador', as_index=False).size().rename(columns={'size': 'Cantidad'}).sort_values('Cantidad', ascending=False).reset_index(drop=True)

                wb2 = load_workbook(output_buffer)
                ws2 = wb2.create_sheet('Ausencias')
                for ci, n in enumerate(['Identificador', 'Nombre', 'Fecha', 'Novedad'], start=1):
                    c = ws2.cell(row=1, column=ci, value=n)
                    c.font, c.fill, c.alignment, c.border = Font(name='Arial', bold=True, color='FFFFFF', size=10), PatternFill('solid', fgColor='1F4E79'), Alignment(horizontal='center', vertical='center'), brd
                for ri, row in listado.iterrows():
                    for ci, val in enumerate(row, start=1):
                        c = ws2.cell(row=ri + 2, column=ci, value=val)
                        c.font, c.fill, c.alignment, c.border = Font(name='Arial', size=9), PatternFill('solid', fgColor='FFFFFF' if ri % 2 == 0 else 'F2F2F2'), Alignment(horizontal='left' if ci <= 2 else 'center', vertical='center'), brd
                ws2.column_dimensions['A'].width, ws2.column_dimensions['B'].width, ws2.column_dimensions['C'].width, ws2.column_dimensions['D'].width = 15, 35, 18, 30

                for ci, n in enumerate(['Identificador', 'Cantidad'], start=9):
                    c = ws2.cell(row=1, column=ci, value=n)
                    c.font, c.fill, c.alignment, c.border = Font(name='Arial', bold=True, color='FFFFFF', size=10), PatternFill('solid', fgColor='1F4E79'), Alignment(horizontal='center', vertical='center'), brd
                for ri, row in resumen.iterrows():
                    c_id = ws2.cell(row=ri + 2, column=9, value=row['Identificador'])
                    c_id.font, c_id.fill, c_id.alignment, c_id.border = Font(name='Arial', size=9), PatternFill('solid', fgColor='FFFFFF' if ri % 2 == 0 else 'F2F2F2'), Alignment(horizontal='left', vertical='center'), brd
                    c_cnt = ws2.cell(row=ri + 2, column=10, value=row['Cantidad'])
                    c_cnt.font, c_cnt.fill, c_cnt.alignment, c_cnt.border = Font(name='Arial', size=9), PatternFill('solid', fgColor='FFFFFF' if ri % 2 == 0 else 'F2F2F2'), Alignment(horizontal='center', vertical='center'), brd
                ws2.column_dimensions['I'].width, ws2.column_dimensions['J'].width = 15, 12
                ws2.freeze_panes, ws2.auto_filter.ref = 'A2', f"A1:J{len(listado) + 1}"
                
                output_buffer = io.BytesIO()
                wb2.save(output_buffer)

                # ── Módulo de Consolidación Multi-Periodo HTCC ──────────────────
                output_buffer.seek(0)
                df_rep = pd.read_excel(output_buffer, sheet_name='Reporte_Horizontal')
                COLS_FIJAS_REP = ["Identificador", "Concepto"]
                cols_fecha_rep = [c for c in df_rep.columns if c not in COLS_FIJAS_REP]
                
                df_long = df_rep.melt(id_vars=COLS_FIJAS_REP, value_vars=cols_fecha_rep, var_name="FechaCorta", value_name="Valor")
                
                def fecha_corta_a_dt(texto):
                    meses_inv = {v: k for k, v in MESES_ES.items()}
                    try:
                        partes = str(texto).split("-")
                        return pd.Timestamp(2026, meses_inv[partes[1].lower()], int(partes[0]))
                    except: return pd.NaT

                df_long["FechaReal"] = df_long["FechaCorta"].apply(fecha_corta_a_dt)
                df_long = df_long.dropna(subset=["FechaReal"])
                df_long = df_long[df_long["Valor"].notna() & df_long["Valor"].astype(str).str.strip().ne("") & df_long["Valor"].astype(str).str.strip().ne("nan")]
                df_long["Identificador"] = df_long["Identificador"].apply(limpiar_id_a_texto)

                def asignar_periodo(fecha):
                    for per, (inicio, fin) in PERIODOS.items():
                        if inicio <= fecha <= fin: return per
                    return None

                df_long["Periodo"] = df_long["FechaReal"].apply(asignar_periodo)
                df_long = df_long[df_long["Periodo"].notna()]

                archivo_htcc.seek(0)
                wb_htcc = load_workbook(archivo_htcc)
                ws_htcc = wb_htcc[HOJA_LIBRO3]
                
                FILA_ENCABEZADO = 4
                COL_INICIO_FECHAS = 12

                col_periodo_libro3 = col_id_libro3 = col_concepto_libro3 = None
                for cell in ws_htcc[FILA_ENCABEZADO]:
                    if cell.value is None: continue
                    val = str(cell.value).strip().lower()
                    if val == "periodo": col_periodo_libro3 = cell.column
                    if "identificador" in val: col_id_libro3 = cell.column
                    if "nombre concepto" in val: col_concepto_libro3 = cell.column

                col_fecha_libro3 = {}
                max_col_fecha = COL_INICIO_FECHAS
                for col_idx in range(COL_INICIO_FECHAS, ws_htcc.max_column + 1):
                    val = ws_htcc.cell(row=FILA_ENCABEZADO, column=col_idx).value
                    if val is None: continue
                    try:
                        fecha = pd.Timestamp(val).normalize()
                        col_fecha_libro3[fecha] = col_idx
                        if col_idx > max_col_fecha: max_col_fecha = col_idx
                    except: pass

                COL_TOTAL_IDX = max_col_fecha + 1
                COL_DIF_IDX = COL_TOTAL_IDX + 1

                indice_filas = {}
                for row in ws_htcc.iter_rows(min_row=FILA_ENCABEZADO + 1):
                    periodo_val = id_val = conc_val = None
                    fila_num = row[0].row
                    for cell in row:
                        if cell.column == col_periodo_libro3: periodo_val = str(cell.value).strip().upper() if cell.value else None
                        if cell.column == col_id_libro3: id_val = limpiar_id_a_texto(cell.value)
                        if cell.column == col_concepto_libro3: conc_val = str(cell.value).strip().lower() if cell.value else None
                    if periodo_val and id_val and conc_val and id_val not in ("None", "nan", ""):
                        indice_filas[(periodo_val, id_val, conc_val)] = fila_num

                for _, fila_long in df_long.iterrows():
                    id_str = fila_long["Identificador"]
                    concepto_long = str(fila_long["Concepto"]).strip()
                    fecha_real_long = fila_long["FechaReal"].normalize()
                    valor_raw = fila_long["Valor"]
                    periodo_long = str(fila_long["Periodo"]).strip().upper()

                    concepto_libro3 = MAPA_CONCEPTOS.get(concepto_long)
                    if concepto_libro3 is None: continue

                    clave = (periodo_long, id_str, concepto_libro3)
                    fila_excel = indice_filas.get(clave)
                    if fila_excel is None: continue

                    col_excel = col_fecha_libro3.get(fecha_real_long)
                    if col_excel is None: continue

                    try:
                        valor_final = round(float(str(valor_raw)), 2)
                        es_numero = True
                    except:
                        valor_final = str(valor_raw).strip().upper()
                        es_numero = False

                    cell = ws_htcc.cell(row=fila_excel, column=col_excel, value=valor_final)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if es_numero:
                        cell.number_format = '#,##0.00'
                        cell.font = Font(name="Arial", size=9)
                        cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
                    else:
                        colores = COLORES_LETRAS.get(valor_final, {"bg":"FFFFFF","fg":"000000"})
                        cell.font = Font(name="Arial", size=9, bold=True, color=colores["fg"])
                        cell.fill = PatternFill(fill_type="solid", fgColor=colores["bg"])

                cols_por_periodo = {p.upper(): [] for p in PERIODOS}
                for f_dt, c_idx in col_fecha_libro3.items():
                    for per, (inicio, fin) in PERIODOS.items():
                        if inicio.normalize() <= f_dt <= fin.normalize():
                            cols_por_periodo[per.upper()].append(c_idx)
                            break

                ws_htcc.cell(row=FILA_ENCABEZADO, column=COL_TOTAL_IDX, value="Total").font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                ws_htcc.cell(row=FILA_ENCABEZADO, column=COL_TOTAL_IDX).fill = PatternFill(fill_type="solid", fgColor="000000")
                ws_htcc.cell(row=FILA_ENCABEZADO, column=COL_TOTAL_IDX).alignment = Alignment(horizontal="center", vertical="center")

                ws_htcc.cell(row=FILA_ENCABEZADO, column=COL_DIF_IDX, value="Diferencia").font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                ws_htcc.cell(row=FILA_ENCABEZADO, column=COL_DIF_IDX).fill = PatternFill(fill_type="solid", fgColor="000000")
                ws_htcc.cell(row=FILA_ENCABEZADO, column=COL_DIF_IDX).alignment = Alignment(horizontal="center", vertical="center")

                for (periodo, id_str, conc_libro3), fila_excel in indice_filas.items():
                    cols_periodo = cols_por_periodo.get(periodo, [])
                    if not cols_periodo: continue
                    col_ini = get_column_letter(min(cols_periodo))
                    col_fin = get_column_letter(max(cols_periodo))
                    
                    formula_total = f"=SUM({col_ini}{fila_excel}:{col_fin}{fila_excel})"
                    color_fondo = "FFFFFF"
                    if conc_libro3 == "recargo nocturno 0.35%": color_fondo = "F2F2F2"
                    elif conc_libro3 == "recargo dominical compensado": color_fondo = "E2EFDA"
                    elif conc_libro3 == "recargo festivo": color_fondo = "FFF2CC"

                    c_total = ws_htcc.cell(row=fila_excel, column=COL_TOTAL_IDX, value=formula_total)
                    c_total.number_format, c_total.font, c_total.alignment, c_total.fill = '#,##0.00', Font(name="Arial", size=9, bold=True), Alignment(horizontal="center", vertical="center"), PatternFill(fill_type="solid", fgColor=color_fondo)

                    formula_dif = f"=ROUND({get_column_letter(7)}{fila_excel}-{get_column_letter(COL_TOTAL_IDX)}{fila_excel},0)"
                    c_dif = ws_htcc.cell(row=fila_excel, column=COL_DIF_IDX, value=formula_dif)
                    c_dif.number_format, c_dif.font, c_dif.alignment, c_dif.fill = '#,##0.00', Font(name="Arial", size=9, bold=True), Alignment(horizontal="center", vertical="center"), PatternFill(fill_type="solid", fgColor="FFFFFF")

                htcc_buffer = io.BytesIO()
                wb_htcc.save(htcc_buffer)

                # ── Hoja Análisis C ───────────────────────────────────────────────
                fecha_inicio_sem = pd.Timestamp(fecha_inicio_input)
                wb_c = load_workbook(output_buffer)
                ws_c = wb_c['Reporte_Horizontal']
                
                actual_col_inicio = 3
                for col in range(1, ws_c.max_column + 1):
                    val_header = str(ws_c.cell(row=1, column=col).value or "")
                    if "-" in val_header and any(m in val_header.lower() for m in ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]):
                        actual_col_inicio = col
                        break

                columnas_c = []
                MESES_ES_INV = {v: k for k, v in MESES_ES.items()}
                for col in range(actual_col_inicio, ws_c.max_column + 1):
                    fecha_str = ws_c.cell(row=1, column=col).value
                    dia_str   = ws_c.cell(row=2, column=col).value
                    if fecha_str is None: continue
                    try:
                        pf = str(fecha_str).strip().split("-")
                        fecha_dt = pd.Timestamp(fecha_inicio_sem.year, MESES_ES_INV[pf[1].lower()], int(pf[0]))
                        columnas_c.append({"col": col, "fecha_str": fecha_str, "dia": str(dia_str).strip().lower().split()[0] if dia_str else "", "fecha_dt": fecha_dt})
                    except: pass

                columnas_c = [c for c in columnas_c if c["fecha_dt"] >= fecha_inicio_sem]
                semanas_c = []
                fecha_sem = fecha_inicio_sem
                num_sem = 1
                while columnas_c and fecha_sem <= columnas_c[-1]["fecha_dt"]:
                    fecha_fin_sem = fecha_sem + pd.Timedelta(days=6)
                    cols_semana = [c for c in columnas_c if fecha_sem <= c["fecha_dt"] <= fecha_fin_sem]
                    if cols_semana:
                        semanas_c.append({"num": num_sem, "label": f"Sem{num_sem}", "inicio": fecha_sem, "fin": fecha_fin_sem, "columnas": cols_semana})
                    fecha_sem = fecha_fin_sem + pd.Timedelta(days=1)
                    num_sem += 1

                df_id_clean = pd.to_numeric(df[col_id], errors='coerce').fillna(0).astype(int).astype(str)
                mapa_nombres_c = dict(zip(df_id_clean, df['Nombre'].astype(str).str.strip()))

                empleados_c = []
                for row in range(3, ws_c.max_row + 1):
                    id_val = ws_c.cell(row=row, column=1).value
                    conc_val = None
                    for c_idx in range(2, actual_col_inicio):
                        c_val = str(ws_c.cell(row=row, column=c_idx).value).strip()
                        if c_val == "HT Normales":
                            conc_val = c_val
                            break
                    if id_val is None or conc_val is None: continue
                    try: id_clean = str(int(float(id_val)))
                    except: id_clean = str(id_val).strip()
                    
                    valores = {col_info["col"]: str(ws_c.cell(row=row, column=col_info["col"]).value).strip() if ws_c.cell(row=row, column=col_info["col"]).value is not None else None for col_info in columnas_c}
                    empleados_c.append({"id": id_clean, "nombre": mapa_nombres_c.get(id_clean, "No Encontrado"), "concepto": conc_val, "fila": row, "valores": valores})

                resultados_c = []
                max_fechas_c = 0
                for emp in empleados_c:
                    for semana in semanas_c:
                        cols_semana = semana["columnas"]
                        cols_domingo = [d for d in cols_semana if d["dia"] == "domingo"]
                        cols_c_val = [d for d in cols_semana if emp["valores"].get(d["col"]) == "C"]
                        trabajo_domingo = any(emp["valores"].get(d["col"]) not in (None, "A","V","S","INC","REN","LR","PD") for d in cols_domingo)
                        cantidad_c = len(cols_c_val)
                        
                        if len(cols_domingo) == 0: estado = "-" if cantidad_c == 0 else "NO CUMPLE"
                        elif trabajo_domingo: estado = "CUMPLE" if cantidad_c == 1 else "-" if cantidad_c == 0 else "NO CUMPLE"
                        else: estado = "-" if cantidad_c == 0 else "NO CUMPLE"

                        if estado != "-" and cantidad_c > 0:
                            fechas_c_list = [c["fecha_str"] for c in cols_c_val]
                            max_fechas_c = max(max_fechas_c, len(fechas_c_list))
                            fila = {"Identificador": emp["id"], "Nombre": emp["nombre"], "Concepto": emp["concepto"], "Trabajo Domingo": "SI" if trabajo_domingo else "NO", "Cant. C": cantidad_c, "Estado": estado}
                            for i, f in enumerate(fechas_c_list, 1): fila[f"Fecha {i}"] = f
                            resultados_c.append(fila)

                df_c = pd.DataFrame(resultados_c) if resultados_c else pd.DataFrame()

                if not df_c.empty:
                    ws_out_c = wb_c.create_sheet("Analisis C")
                    headers_base = ["Identificador", "Nombre", "Concepto", "Trabajo Domingo", "Cant. C", "Estado"]
                    headers_fechas = [f"Fecha {i}" for i in range(1, max_fechas_c + 1)]
                    headers_c = headers_base + headers_fechas
                    brd_c = Border(left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"), top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))

                    for col_i, h in enumerate(headers_c, 1):
                        cell = ws_out_c.cell(row=1, column=col_i, value=h)
                        cell.font, cell.fill, cell.alignment, cell.border = Font(name="Arial", size=10, bold=True, color="FFFFFF"), PatternFill("solid", fgColor='1F4E79'), Alignment(horizontal="center", vertical="center"), brd_c

                    n_cumple = len(df_c[df_c["Estado"] == "CUMPLE"])
                    n_no_cumple = len(df_c[df_c["Estado"] == "NO CUMPLE"])

                    for row_i, (_, row) in enumerate(df_c.iterrows(), 2):
                        color_fondo = 'D9E1F2' if row_i % 2 == 0 else 'FFFFFF'
                        for col_i, h in enumerate(headers_c, 1):
                            val = row.get(h, "")
                            cell = ws_out_c.cell(row=row_i, column=col_i, value=val)
                            cell.alignment = Alignment(horizontal="left" if h == "Nombre" else "center", vertical="center")
                            cell.border = brd_c
                            if h == "Estado":
                                cell.fill = PatternFill("solid", fgColor='92D050' if val == "CUMPLE" else 'FF0000')
                                cell.font = Font(name="Arial", size=9, bold=True, color='000000' if val == "CUMPLE" else 'FFFFFF')
                            elif h.startswith("Fecha ") and val:
                                cell.fill, cell.font = PatternFill("solid", fgColor='E2EFDA'), Font(name="Arial", size=9, bold=True, color="1A5C2A")
                            else:
                                cell.fill, cell.font = PatternFill("solid", fgColor=color_fondo), Font(name="Arial", size=9)

                    for col_i, ancho in enumerate([15, 35, 20, 16, 10, 12] + [12] * max_fechas_c, 1):
                        ws_out_c.column_dimensions[get_column_letter(col_i)].width = ancho
                    
                    fila_total_c = len(df_c) + 2
                    ws_out_c.cell(row=fila_total_c, column=1, value="TOTALES").font = Font(bold=True)
                    ws_out_c.cell(row=fila_total_c, column=5, value=f"CUMPLE: {n_cumple}").font = Font(bold=True, color="2E7D32")
                    ws_out_c.cell(row=fila_total_c, column=6, value=f"NO CUMPLE: {n_no_cumple}").font = Font(bold=True, color="C62828")
                    ws_out_c.freeze_panes, ws_out_c.auto_filter.ref = 'A2', f"A1:{get_column_letter(len(headers_c))}{fila_total_c - 1}"
                
                output_buffer = io.BytesIO()
                wb_c.save(output_buffer)
                
                st.session_state.output_bytes = output_buffer.getvalue()
                st.session_state.htcc_bytes = htcc_buffer.getvalue()
                st.session_state.listado = listado
                st.session_state.resumen = resumen
                st.session_state.df_c = df_c
                st.session_state.resultados_c = resultados_c
                st.session_state.procesado = True

            except Exception as e:
                st.error(f"❌ Ocurrió un error inesperado al procesar: {e}")

    # ── BLOQUE DE RENDERIZADO VISUAL FUERA DEL BOTÓN ──────────────────────────
    if st.session_state.procesado:
        st.success("🎉 ¡Reporte y Consolidación Multi-Periodo procesados exitosamente!")
        
        col_down1, col_down2 = st.columns(2)
        with col_down1:
            st.download_button(
                label="📥 Descargar Reporte Horizontal Procesado",
                data=st.session_state.output_bytes,
                file_name="Reporte_Horizontal_Asistencia.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with col_down2:
            st.download_button(
                label="📥 Descargar Plantilla HTCC Consolidada",
                data=st.session_state.htcc_bytes,
                file_name="HTCC_Consolidado_2026.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.markdown("---")
        st.header("📋 Vista Previa de Resultados")
        
        tab_aus, tab_comp = st.tabs([
            "📄 Hoja de Ausencias", 
            "🔍 Análisis de Compensatorios"
        ])

        with tab_aus:
            col_titulo, col_metrica = st.columns([3, 1])
            with col_titulo:
                st.subheader("Registros Detallados de Ausencias")
            with col_metrica:
                total_ausencias = len(st.session_state.listado)
                st.markdown(f"""
                    <div style="background-color:#FFEB9C; padding:5px 15px; border-radius:15px; text-align:center; border:1px solid #FFC7CE; margin-top:5px;">
                        <strong style="color:#9C0006; font-size:16px;">Total: {total_ausencias}</strong>
                    </div>
                """, unsafe_allow_html=True)
            st.dataframe(st.session_state.listado, use_container_width=True, hide_index=True)
            st.subheader("Resumen Consolidado por Persona")
            st.dataframe(st.session_state.resumen, use_container_width=True, hide_index=True)
            
        with tab_comp:
            st.subheader("Validación de Compensatorios (Analisis C)")
            if st.session_state.resultados_c:
                st.write("Filtrar por Estado:")
                col_cumple, col_nocumple, _ = st.columns([1, 1, 3])
                with col_cumple:
                    chk_cumple = st.checkbox("CUMPLE", value=True, key="filtro_cumple")
                with col_nocumple:
                    chk_nocumple = st.checkbox("NO CUMPLE", value=True, key="filtro_nocumple")
                
                estados_activos = []
                if chk_cumple:
                    estados_activos.append("CUMPLE")
                if chk_nocumple:
                    estados_activos.append("NO CUMPLE")
                
                df_c_filtrado = st.session_state.df_c[st.session_state.df_c["Estado"].isin(estados_activos)]
                st.dataframe(df_c_filtrado, use_container_width=True, hide_index=True)
            else:
                st.info("No se encontraron registros de compensatorios que requieran validación para el periodo seleccionado.")
